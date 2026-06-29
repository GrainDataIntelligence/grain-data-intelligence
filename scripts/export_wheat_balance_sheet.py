import json
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
SOURCE = Path(r"C:\Users\Myburgh Swiegers\Downloads\Wheat_Data_Table_20260626.xlsx")
LOCAL_OUTPUT = ROOT / "data" / "wheat_balance_sheet.json"
PUBLIC_OUTPUT = ROOT / "charting-react" / "public" / "data" / "balance_sheet" / "wheat.json"
OFFICIAL_OUTPUT = (
    Path(r"C:\Users\Myburgh Swiegers\Projects\GrainDataIntelligence\grain-data-intelligence")
    / "public"
    / "data"
    / "balance_sheet"
    / "wheat.json"
)

MONTHS = ["Oct", "Nov", "Dec", "Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep"]
MONTH_ORDER = {month: index + 1 for index, month in enumerate(MONTHS)}

METRICS = [
    {
        "key": "producerDeliveries",
        "label": "Producer Deliveries",
        "type": "single",
        "items": ["Deliveries directly from farms"],
    },
    {
        "key": "imports",
        "label": "Imports",
        "type": "single",
        "items": ["Imports destined for RSA"],
    },
    {
        "key": "processedLocalMarket",
        "label": "Processed for the Local Market",
        "type": "sum",
        "items": ["Human consumption", "Animal feed", "Gristing", "Bio-fuel"],
    },
    {
        "key": "exports",
        "label": "Exports",
        "type": "sum",
        "items": [
            "RSA Export Harbours",
            "RSA Export Border posts",
            "Product Export African countries",
            "Product Export Other countries",
            "Total RSA Export",
            "Total Product Export",
        ],
    },
    {
        "key": "totalDemand",
        "label": "Total Demand",
        "type": "sum",
        "items": [
            "Human consumption",
            "Animal feed",
            "Gristing",
            "Bio-fuel",
            "RSA Export Harbours",
            "RSA Export Border posts",
            "Product Export African countries",
            "Product Export Other countries",
            "Total RSA Export",
            "Total Product Export",
            "Seed for planting purposes",
            "Released to end-consumer(s)",
            "Withdrawn by producers",
        ],
    },
    {
        "key": "endingStocks",
        "label": "Ending Stocks",
        "type": "stock",
        "items": ["Unutilised Closing stock"],
    },
]


def month_info(value):
    if isinstance(value, datetime):
        month = value.strftime("%b")
        year = value.year
    else:
        parts = str(value).strip().split()
        month = parts[0][:3]
        year = int(parts[1]) if len(parts) > 1 and parts[1].isdigit() else None
    return month, year, MONTH_ORDER.get(month)


def number(value):
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace(",", "").replace(" ", "").strip()
    if not text:
        return 0.0
    return float(text)


def build_payload(source):
    wb = openpyxl.load_workbook(source, read_only=True, data_only=True)
    ws = wb["Data Table Wheat"]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    index = {header: position for position, header in enumerate(headers)}

    raw = defaultdict(float)
    publication_dates = set()
    row_count = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[index["Marketing Year"]] is None or row[index["Return Item"]] is None:
            continue

        marketing_year = str(row[index["Marketing Year"]]).strip()
        status = str(row[index["Status"]]).strip()
        publication_date = row[index["Publication Date"]]
        return_item = str(row[index["Return Item"]]).strip()
        month, calendar_year, order = month_info(row[index["Return Month"]])

        if not marketing_year or order is None:
            continue

        if isinstance(publication_date, datetime):
            publication_dates.add(publication_date.strftime("%Y-%m-%d"))
        elif publication_date:
            publication_dates.add(str(publication_date))

        key = (marketing_year, status, month, calendar_year, order, return_item)
        raw[key] += number(row[index["Tonnage"]])
        row_count += 1

    metric_rows = []
    years = sorted({key[0] for key in raw.keys()})

    for marketing_year in years:
        for month in MONTHS:
            order = MONTH_ORDER[month]
            candidates = [key for key in raw.keys() if key[0] == marketing_year and key[2] == month and key[4] == order]
            if not candidates:
                continue
            status = candidates[0][1]
            calendar_year = candidates[0][3]
            month_values = {key[5]: raw[key] for key in candidates}

            for metric in METRICS:
                monthly_value = sum(month_values.get(item, 0.0) for item in metric["items"])
                if monthly_value == 0 and metric["type"] != "stock":
                    continue
                metric_rows.append(
                    {
                        "marketingYear": marketing_year,
                        "status": status,
                        "month": month,
                        "calendarYear": calendar_year,
                        "monthOrder": order,
                        "metric": metric["key"],
                        "value": round(monthly_value, 2),
                    }
                )

    payload = {
        "generatedAt": datetime.now().isoformat(timespec="seconds"),
        "sourceFile": str(source),
        "rowCount": row_count,
        "commodity": "Wheat",
        "months": MONTHS,
        "metrics": [{"key": metric["key"], "label": metric["label"], "type": metric["type"]} for metric in METRICS],
        "years": years,
        "publicationDates": sorted(publication_dates),
        "rows": metric_rows,
    }
    return payload


def write_json(path, payload):
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2), encoding="utf-8")


def main():
    payload = build_payload(SOURCE)
    write_json(LOCAL_OUTPUT, payload)
    write_json(PUBLIC_OUTPUT, payload)
    if OFFICIAL_OUTPUT.parents[2].exists():
        write_json(OFFICIAL_OUTPUT, payload)
    print(f"Wrote {len(payload['rows'])} wheat balance sheet rows")
    print(f"Years: {payload['years'][0]} to {payload['years'][-1]}")
    print(f"Output: {PUBLIC_OUTPUT}")


if __name__ == "__main__":
    main()
