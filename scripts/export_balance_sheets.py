import json
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
PUBLIC_DIR = ROOT / "charting-react" / "public" / "data" / "balance_sheet"
LOCAL_DIR = ROOT / "data" / "balance_sheet"
OFFICIAL_DIR = (
    Path(r"C:\Users\Myburgh Swiegers\Projects\GrainDataIntelligence\grain-data-intelligence")
    / "public"
    / "data"
    / "balance_sheet"
)

EXPORT_COMPONENTS = [
    "RSA Export Harbours",
    "RSA Export Border posts",
    "Product Export African countries",
    "Product Export Other countries",
]
EXPORT_TOTALS = ["Total RSA Export", "Total Product Export"]

CONFIGS = [
    {
        "key": "wheat",
        "commodity": "Wheat",
        "source": Path(r"C:\Users\Myburgh Swiegers\Downloads\Wheat_Data_Table_20260525.xlsx"),
        "sheet": "Data Table Wheat",
        "months": ["Oct", "Nov", "Dec", "Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep"],
        "viewLabels": {"Human": "Wheat"},
        "metrics": "wheat",
    },
    {
        "key": "sunflowers",
        "commodity": "Sunflowers",
        "source": Path(r"C:\Users\Myburgh Swiegers\Downloads\Sunflower_Data_Table_20260525.xlsx"),
        "sheet": "Data Table SUN",
        "months": ["Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "Jan", "Feb"],
        "viewLabels": {"Sunflower": "Sunflowers"},
        "metrics": "oilseeds",
    },
    {
        "key": "soybeans",
        "commodity": "Soybeans",
        "source": Path(r"C:\Users\Myburgh Swiegers\Downloads\Soybeans_Data_Table_20260525.xlsx"),
        "sheet": "Data Table SOA",
        "months": ["Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "Jan", "Feb"],
        "viewLabels": {"Soybeans": "Soybeans"},
        "metrics": "oilseeds",
    },
    {
        "key": "maize",
        "commodity": "Maize",
        "source": Path(r"C:\Users\Myburgh Swiegers\Downloads\Maize_Data_Table_20260525.xlsx"),
        "sheet": "Data Table Maize",
        "months": ["May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "Jan", "Feb", "Mar", "Apr"],
        "viewLabels": {"White": "White Maize", "Yellow": "Yellow Maize"},
        "totalView": "Total Maize",
        "metrics": "maize",
    },
]


METRIC_DEFINITIONS = {
    "wheat": [
        ("producerDeliveries", "Producer Deliveries", "single", ["Deliveries directly from farms"]),
        ("imports", "Imports", "single", ["Imports destined for RSA"]),
        ("processedLocalMarket", "Processed for the Local Market", "sum", ["Human consumption", "Animal feed", "Gristing", "Bio-fuel"]),
        ("exports", "Exports", "exports", []),
        (
            "totalDemand",
            "Total Demand",
            "demand",
            ["Human consumption", "Animal feed", "Gristing", "Bio-fuel", "Seed for planting purposes", "Released to end-consumer(s)", "Withdrawn by producers"],
        ),
        ("endingStocks", "Ending Stocks", "stock", ["Unutilised Closing stock"]),
    ],
    "oilseeds": [
        ("producerDeliveries", "Producer Deliveries", "single", ["Deliveries directly from farms"]),
        ("imports", "Imports", "single", ["Imports destined for RSA"]),
        (
            "processedLocalMarket",
            "Processed for the Local Market",
            "sum",
            ["Human consumption for Local Market", "Animal feed for Local Market", "Oil and oilcake for Local Market"],
        ),
        ("exports", "Exports", "exports", []),
        (
            "totalDemand",
            "Total Demand",
            "demand",
            [
                "Human consumption for Local Market",
                "Animal feed for Local Market",
                "Oil and oilcake for Local Market",
                "Seed for planting purposes",
                "Released to end-consumer(s)",
                "Withdrawn by producers",
            ],
        ),
        ("endingStocks", "Ending Stocks", "stock", ["Unutilised Closing stock"]),
    ],
    "maize": [
        ("producerDeliveries", "Producer Deliveries", "single", ["Deliveries directly from farms"]),
        ("imports", "Imports", "single", ["Imports destined for RSA"]),
        ("processedLocalMarket", "Processed for the Local Market", "sum", ["Human Consumption", "Animal feed / Industrial", "Gristing"]),
        ("exports", "Exports", "exports", []),
        (
            "totalDemand",
            "Total Demand",
            "demand",
            ["Human Consumption", "Animal feed / Industrial", "Gristing", "Released to end-consumer(s)", "Withdrawn by producers"],
        ),
        ("endingStocks", "Ending Stocks", "stock", ["Unutilised Closing stock"]),
    ],
}


def month_info(value, month_order):
    if isinstance(value, datetime):
        month = value.strftime("%b")
        year = value.year
    else:
        parts = str(value).strip().split()
        month = parts[0][:3]
        year = int(parts[1]) if len(parts) > 1 and parts[1].isdigit() else None
    return month, year, month_order.get(month)


def numeric(value):
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace(",", "").replace(" ", "").strip()
    return float(text) if text else 0.0


def metric_value(metric_type, items, month_values):
    if metric_type == "exports":
        total_value = sum(month_values.get(item, 0.0) for item in EXPORT_TOTALS)
        return total_value if total_value else sum(month_values.get(item, 0.0) for item in EXPORT_COMPONENTS)
    if metric_type == "demand":
        local = sum(month_values.get(item, 0.0) for item in items)
        export_total = sum(month_values.get(item, 0.0) for item in EXPORT_TOTALS)
        exports = export_total if export_total else sum(month_values.get(item, 0.0) for item in EXPORT_COMPONENTS)
        return local + exports
    return sum(month_values.get(item, 0.0) for item in items)


def build_payload(config):
    month_order = {month: index + 1 for index, month in enumerate(config["months"])}
    wb = openpyxl.load_workbook(config["source"], read_only=True, data_only=True)
    ws = wb[config["sheet"]]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    index = {header: position for position, header in enumerate(headers) if header}

    raw = defaultdict(float)
    publication_dates = set()
    row_count = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[index["Marketing Year"]] is None or row[index["Return Item"]] is None:
            continue
        sub_cereal = str(row[index["Sub Cereal"]]).strip()
        view = config["viewLabels"].get(sub_cereal, sub_cereal)
        marketing_year = str(row[index["Marketing Year"]]).strip()
        status = str(row[index["Status"]]).strip()
        return_item = str(row[index["Return Item"]]).strip()
        month, calendar_year, order = month_info(row[index["Return Month"]], month_order)

        if not marketing_year or order is None:
            continue

        publication_date = row[index["Publication Date"]]
        if isinstance(publication_date, datetime):
            publication_dates.add(publication_date.strftime("%Y-%m-%d"))
        elif publication_date:
            publication_dates.add(str(publication_date))

        raw[(view, marketing_year, status, month, calendar_year, order, return_item)] += numeric(row[index["Tonnage"]])
        row_count += 1

    metric_defs = METRIC_DEFINITIONS[config["metrics"]]
    rows = []
    years = sorted({key[1] for key in raw.keys()})
    views = list(dict.fromkeys(config["viewLabels"].values()))

    for view in views:
        for marketing_year in years:
            for month in config["months"]:
                order = month_order[month]
                candidates = [key for key in raw.keys() if key[0] == view and key[1] == marketing_year and key[3] == month and key[5] == order]
                if not candidates:
                    continue
                status = candidates[0][2]
                calendar_year = candidates[0][4]
                month_values = {key[6]: raw[key] for key in candidates}

                for metric_key, label, metric_type, items in metric_defs:
                    value = metric_value(metric_type, items, month_values)
                    if value == 0 and metric_type != "stock":
                        continue
                    rows.append(
                        {
                            "view": view,
                            "marketingYear": marketing_year,
                            "status": status,
                            "month": month,
                            "calendarYear": calendar_year,
                            "monthOrder": order,
                            "metric": metric_key,
                            "value": round(value, 2),
                        }
                    )

    if config.get("totalView"):
        totals = defaultdict(float)
        metadata = {}
        for row in rows:
            key = (row["marketingYear"], row["month"], row["monthOrder"], row["metric"])
            totals[key] += row["value"]
            metadata[key] = (row["status"], row["calendarYear"])
        for (marketing_year, month, order, metric_key), value in totals.items():
            status, calendar_year = metadata[(marketing_year, month, order, metric_key)]
            rows.append(
                {
                    "view": config["totalView"],
                    "marketingYear": marketing_year,
                    "status": status,
                    "month": month,
                    "calendarYear": calendar_year,
                    "monthOrder": order,
                    "metric": metric_key,
                    "value": round(value, 2),
                }
            )
        views.append(config["totalView"])

    return {
        "generatedAt": datetime.now().isoformat(timespec="seconds"),
        "sourceFile": str(config["source"]),
        "rowCount": row_count,
        "commodity": config["commodity"],
        "months": config["months"],
        "views": views,
        "defaultView": views[-1] if config.get("totalView") else views[0],
        "metrics": [{"key": key, "label": label, "type": metric_type} for key, label, metric_type, _ in metric_defs],
        "years": years,
        "publicationDates": sorted(publication_dates),
        "rows": rows,
    }


def write_json(path, payload):
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2), encoding="utf-8")


def main():
    for config in CONFIGS:
        if config["key"] == "maize":
            continue
        payload = build_payload(config)
        for output_dir in [LOCAL_DIR, PUBLIC_DIR, OFFICIAL_DIR]:
            write_json(output_dir / f"{config['key']}.json", payload)
        print(f"{config['commodity']}: wrote {len(payload['rows'])} rows, {payload['years'][0]} to {payload['years'][-1]}")

    from export_maize_balance_sheet import build_payload as build_maize_payload

    maize_payload = build_maize_payload()
    for output_dir in [LOCAL_DIR, PUBLIC_DIR, OFFICIAL_DIR]:
        write_json(output_dir / "maize.json", maize_payload)
    print(f"Maize: wrote {len(maize_payload['rows'])} rows, {maize_payload['years'][0]} to {maize_payload['years'][-1]}")


if __name__ == "__main__":
    main()
