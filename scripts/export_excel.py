from __future__ import annotations

import json
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "Codex Imp and Exp.xlsx"
OUTFILE = ROOT / "data" / "exports.json"

SHEETS = {
    "White RSA Exports": "White Maize",
    "YELLOW RSA EXPORT": "Yellow Maize",
}

COUNTRY_FIXES = {
    "ANGO LA": "ANGOLA",
    "Other ": "Other",
    "Deapsea": "Deepsea",
}

AGGREGATES = {"BNLS + Moz", "Other", "Deepsea"}
CUMULATIVE_COLUMNS = {
    "BNLS + Moz Cumulative",
    "Other Cumulative",
    "Deapsea Cumulative",
}


def clean_header(value: object) -> str:
    text = "" if value is None else str(value).strip()
    return COUNTRY_FIXES.get(text, text)


def as_number(value: object) -> float:
    if value in (None, ""):
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def normalise() -> dict:
    wb = load_workbook(WORKBOOK, data_only=True, read_only=True)
    rows: list[dict] = []
    destinations: set[str] = set()
    years: set[str] = set()
    commodities: list[str] = []

    for sheet_name, commodity in SHEETS.items():
        ws = wb[sheet_name]
        commodities.append(commodity)
        raw_headers = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
        headers = [clean_header(header) for header in raw_headers]

        year_idx = 0
        week_idx = 1
        week_label_idx = 2

        destination_indexes: dict[str, list[int]] = defaultdict(list)
        for idx, header in enumerate(headers):
            if idx <= week_label_idx or not header or header in CUMULATIVE_COLUMNS:
                continue
            destination_indexes[header].append(idx)

        for excel_row in ws.iter_rows(min_row=2, values_only=True):
            year = excel_row[year_idx]
            week_number = excel_row[week_idx]
            week_label = excel_row[week_label_idx]
            if not year or not week_number:
                continue

            year = str(year).strip()
            week_number = int(week_number)
            if week_number > 52:
                continue
            years.add(year)

            weekly_by_destination: dict[str, float] = {}
            for destination, indexes in destination_indexes.items():
                value = sum(as_number(excel_row[idx]) for idx in indexes)
                weekly_by_destination[destination] = value
                destinations.add(destination)

            total = sum(
                weekly_by_destination.get(group, 0.0)
                for group in ("BNLS + Moz", "Other", "Deepsea")
            )
            weekly_by_destination["Total Exports"] = total
            destinations.add("Total Exports")

            for destination, weekly_tons in weekly_by_destination.items():
                rows.append(
                    {
                        "commodity": commodity,
                        "flow": "Exports",
                        "marketingYear": year,
                        "weekNumber": week_number,
                        "weekLabel": str(week_label or ""),
                        "destination": destination,
                        "weeklyTons": round(weekly_tons, 3),
                        "isAggregate": destination in AGGREGATES or destination == "Total Exports",
                    }
                )

    cumulative: dict[tuple[str, str, str], float] = defaultdict(float)
    rows.sort(key=lambda item: (item["commodity"], item["marketingYear"], item["destination"], item["weekNumber"]))
    for row in rows:
        key = (row["commodity"], row["marketingYear"], row["destination"])
        cumulative[key] += row["weeklyTons"]
        row["cumulativeTons"] = round(cumulative[key], 3)

    return {
        "generatedFrom": WORKBOOK.name,
        "commodities": commodities,
        "flows": ["Exports"],
        "marketingYears": sorted(years),
        "destinations": sorted(destinations, key=lambda value: (value != "Total Exports", value)),
        "rows": rows,
    }


def main() -> None:
    OUTFILE.parent.mkdir(exist_ok=True)
    OUTFILE.write_text(json.dumps(normalise(), indent=2), encoding="utf-8")
    print(f"Wrote {OUTFILE}")


if __name__ == "__main__":
    main()
