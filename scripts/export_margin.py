import json
import csv
from datetime import datetime
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
RAW_DIR = ROOT / "data" / "raw" / "margin"
OUTFILE = ROOT / "charting-react" / "public" / "data" / "im_safex.json"
PRICE_HISTORY_DIR = ROOT / "charting-react" / "public" / "data" / "price_history"

ALLOWED_COMMODITIES = {"WMAZ", "YMAZ", "WEAT", "SUNS", "SOYB", "CORN"}
COMMODITY_ORDER = {"WMAZ": 0, "YMAZ": 1, "WEAT": 2, "SUNS": 3, "SOYB": 4, "CORN": 5}
MAIN_HEDGING_MONTHS = {"Mar", "May", "Jul", "Sep", "Dec"}
MIN_OPEN_INTEREST = 100
PRICE_HISTORY_FILES = {
    "WMAZ": "White_Maize.csv",
    "YMAZ": "Yellow_Maize.csv",
    "WEAT": "Wheat.csv",
    "SUNS": "Sunflower.csv",
    "SOYB": "Soybeans.csv",
    "CORN": "Corn.csv",
}


def as_int(value):
    if value is None:
        return 0
    if isinstance(value, (int, float)):
        return int(round(value))
    text = str(value).replace(" ", "").replace(",", "").strip()
    if not text or text == "-":
        return 0
    return int(round(float(text)))


def as_float(value):
    if value is None:
        return 0
    if isinstance(value, (int, float)):
        return value
    text = str(value).replace(" ", "").replace(",", "").strip()
    if not text or text == "-":
        return 0
    return float(text)


def format_expiry(value):
    if isinstance(value, datetime):
        return value.strftime("%d-%b-%y")
    if value is None:
        return ""
    text = str(value).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d-%b-%y", "%d-%b-%Y"):
        try:
            return datetime.strptime(text, fmt).strftime("%d-%b-%y")
        except ValueError:
            continue
    return text


def expiry_sort_key(expiry):
    try:
        return datetime.strptime(expiry, "%d-%b-%y")
    except ValueError:
        return datetime.max


def expiry_contract_key(expiry):
    try:
        date = datetime.strptime(expiry, "%d-%b-%y")
    except ValueError:
        return None
    return date.strftime("%b"), date.year


def load_latest_open_interest():
    open_interest = {}

    for commodity, file_name in PRICE_HISTORY_FILES.items():
        path = PRICE_HISTORY_DIR / file_name
        if not path.exists():
            continue

        latest_by_contract = {}
        with path.open(newline="", encoding="utf-8-sig") as handle:
            for row in csv.DictReader(handle):
                if row.get("commodity") != commodity:
                    continue
                month = (row.get("contract_month") or "").strip()
                if month not in MAIN_HEDGING_MONTHS:
                    continue
                try:
                    year = int(float(row.get("contract_year") or 0))
                    report_date = datetime.strptime((row.get("date") or "").strip(), "%Y/%m/%d")
                    oi = as_int(row.get("open_interest"))
                except (TypeError, ValueError):
                    continue

                contract_key = (commodity, month, year)
                existing = latest_by_contract.get(contract_key)
                if existing is None or report_date > existing[0]:
                    latest_by_contract[contract_key] = (report_date, oi)

        for contract_key, (_, oi) in latest_by_contract.items():
            open_interest[contract_key] = oi

    return open_interest


def latest_source():
    files = sorted(RAW_DIR.glob("APD IM Parameters*.xlsx"), key=lambda path: path.stat().st_mtime, reverse=True)
    if not files:
        raise FileNotFoundError(f"No APD IM Parameters Excel files found in {RAW_DIR}")
    return files[0]


def export():
    source = latest_source()
    latest_open_interest = load_latest_open_interest()
    workbook = openpyxl.load_workbook(source, data_only=True)
    sheet = workbook.active

    headers = [sheet.cell(row=4, column=column).value for column in range(1, sheet.max_column + 1)]
    header_map = {str(value).strip(): index + 1 for index, value in enumerate(headers) if value}

    rows = []
    for row in range(5, sheet.max_row + 1):
        commodity = str(sheet.cell(row=row, column=header_map["ShortName"]).value or "").strip()
        if commodity not in ALLOWED_COMMODITIES:
            continue

        expiry = format_expiry(sheet.cell(row=row, column=header_map["ExpiryDate"]).value)
        if not expiry:
            continue
        if expiry.split("-")[1] not in MAIN_HEDGING_MONTHS:
            continue
        contract_key = expiry_contract_key(expiry)
        if contract_key is None:
            continue
        latest_oi = latest_open_interest.get((commodity, contract_key[0], contract_key[1]), 0)
        if latest_oi <= MIN_OPEN_INTEREST:
            continue

        rows.append(
            {
                "commodity": commodity,
                "expiry_date": expiry,
                "open_interest": latest_oi,
                "imr": as_int(sheet.cell(row=row, column=header_map["IMR"]).value),
                "csmr": as_int(sheet.cell(row=row, column=header_map["CSMR"]).value),
                "vsr": as_float(sheet.cell(row=row, column=header_map["VSR"]).value),
                "ssmr": as_int(sheet.cell(row=row, column=header_map["SSMR"]).value),
                "ssg": str(sheet.cell(row=row, column=header_map["SSG"]).value or "").strip(),
            }
        )

    rows.sort(key=lambda item: (COMMODITY_ORDER[item["commodity"]], expiry_sort_key(item["expiry_date"])))

    OUTFILE.parent.mkdir(parents=True, exist_ok=True)
    OUTFILE.write_text(json.dumps(rows, indent=2), encoding="utf-8")
    print(f"Exported {len(rows)} margin rows from {source.name} to {OUTFILE}")


if __name__ == "__main__":
    export()
