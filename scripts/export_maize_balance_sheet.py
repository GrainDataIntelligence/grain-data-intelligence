import json
import re
from datetime import datetime
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
RAW_DIR = ROOT / "data" / "raw" / "balance_sheet"
SOURCE_SHEET = "Maize"
LOCAL_OUTPUT = ROOT / "data" / "balance_sheet" / "maize.json"
PUBLIC_OUTPUT = ROOT / "charting-react" / "public" / "data" / "balance_sheet" / "maize.json"

MONTHS = ["May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "Jan", "Feb", "Mar", "Apr"]
MONTH_ORDER = {month: index + 1 for index, month in enumerate(MONTHS)}
VIEWS = {"White": "White Maize", "Yellow": "Yellow Maize", "Total": "Total Maize"}

ROW_MAP = {
    6: ("openingStock", "Opening Stock", "stock"),
    7: ("producerDeliveries", "Producer Deliveries", "flow"),
    8: ("imports", "Imports", "flow"),
    9: ("surplus", "Surplus", "flow"),
    10: ("totalSupply", "Total Supply", "stock"),
    13: ("processedLocalMarket", "Processed for the Local Market", "flow"),
    14: ("human", "Human", "flow"),
    15: ("animal", "Animal", "flow"),
    16: ("gristing", "Gristing", "flow"),
    17: ("bioFuel", "Bio-Fuel", "flow"),
    18: ("withdrawnByProducers", "Withdrawn by Producers", "flow"),
    19: ("releasedToEndConsumers", "Released to End-Consumers", "flow"),
    20: ("netDispatchesReceipts", "Net Dispatches / Receipts", "flow"),
    21: ("deficit", "Deficit", "flow"),
    22: ("localDemand", "Local Demand", "flow"),
    24: ("exports", "Exports", "flow"),
    25: ("products", "Products", "flow"),
    26: ("africanCountries", "African Countries", "flow"),
    27: ("otherCountries", "Other Countries", "flow"),
    28: ("wholeMaize", "Whole Maize", "flow"),
    29: ("borderPosts", "Border Posts", "flow"),
    30: ("harbours", "Harbours", "flow"),
    32: ("totalDemand", "Total Demand", "flow"),
    34: ("unutilizedClosingStock", "Unutilized Closing Stock", "stock"),
    35: ("storersAndTraders", "Storers and Traders", "stock"),
    36: ("processors", "Processors", "stock"),
}


def latest_source():
    source_files = sorted(
        RAW_DIR.glob("*.xlsx"),
        key=lambda path: path.stat().st_mtime,
        reverse=True,
    )
    for source_file in source_files:
        try:
            workbook = openpyxl.load_workbook(source_file, read_only=True, data_only=True)
        except (OSError, ValueError):
            continue
        try:
            if SOURCE_SHEET in workbook.sheetnames:
                return source_file
        finally:
            workbook.close()

    raise FileNotFoundError(
        f"No .xlsx workbook containing a '{SOURCE_SHEET}' worksheet found in {RAW_DIR}"
    )


def parse_month(value):
    if isinstance(value, datetime):
        return value.year, value.strftime("%b")
    if value is None:
        return None, None
    text = str(value).strip()
    if not text:
        return None, None
    for fmt in ("%Y/%m", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            parsed = datetime.strptime(text, fmt)
            return parsed.year, parsed.strftime("%b")
        except ValueError:
            pass
    return None, None


def marketing_year(calendar_year, month):
    start_year = calendar_year if MONTH_ORDER[month] <= 8 else calendar_year - 1
    return f"{start_year}/{str(start_year + 1)[-2:]}"


def numeric(value):
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace(",", "").replace(" ", "").strip()
    return float(text) if text else 0.0


def publication_date(source):
    date_matches = re.findall(r"(20\d{6})", source.stem)
    if not date_matches:
        raise ValueError(
            f"Could not determine publication date from maize workbook name: {source.name}"
        )
    return datetime.strptime(date_matches[-1], "%Y%m%d").strftime("%Y-%m-%d")


def build_payload(source=None):
    source = source or latest_source()
    wb = openpyxl.load_workbook(source, read_only=False, data_only=True)
    ws = wb[SOURCE_SHEET]

    blocks = []
    column = 2
    while column <= ws.max_column - 2:
        labels = [str(ws.cell(4, column + offset).value or "").strip() for offset in range(3)]
        if labels == ["White", "Yellow", "Total"]:
            year, month = parse_month(ws.cell(3, column + 1).value)
            if year and month in MONTH_ORDER:
                blocks.append((column, year, month))
            column += 3
        else:
            column += 1

    rows = []
    years = set()

    for start_column, calendar_year, month in blocks:
        year_label = marketing_year(calendar_year, month)
        years.add(year_label)
        for row_number, (metric_key, _label, metric_type) in ROW_MAP.items():
            for offset, source_view in enumerate(["White", "Yellow", "Total"]):
                rows.append(
                    {
                        "view": VIEWS[source_view],
                        "marketingYear": year_label,
                        "status": "Final",
                        "month": month,
                        "calendarYear": calendar_year,
                        "monthOrder": MONTH_ORDER[month],
                        "metric": metric_key,
                        "value": round(numeric(ws.cell(row_number, start_column + offset).value), 2),
                    }
                )

    return {
        "generatedAt": datetime.now().isoformat(timespec="seconds"),
        "sourceFile": str(source.relative_to(ROOT)),
        "rowCount": len(rows),
        "commodity": "Maize",
        "months": MONTHS,
        "views": ["White Maize", "Yellow Maize", "Total Maize"],
        "defaultView": "Total Maize",
        "metrics": [
            {"key": key, "label": label, "type": metric_type}
            for key, label, metric_type in ROW_MAP.values()
        ],
        "years": sorted(years),
        "publicationDates": [publication_date(source)],
        "rows": rows,
    }


def write_json(path, payload):
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2), encoding="utf-8")


def main():
    payload = build_payload()
    write_json(LOCAL_OUTPUT, payload)
    write_json(PUBLIC_OUTPUT, payload)
    print(f"Wrote {payload['rowCount']} Maize balance sheet rows")
    print(f"Years: {payload['years'][0]} to {payload['years'][-1]}")
    print(f"Output: {PUBLIC_OUTPUT}")


if __name__ == "__main__":
    main()
