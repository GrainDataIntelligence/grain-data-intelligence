from __future__ import annotations

import json
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "Clean_deliveries_format_maize.xlsx"
OILSEEDS_WORKBOOK = ROOT / "Codex Oilseeds.xlsx"
OUTFILE = ROOT / "data" / "deliveries.json"
PUBLIC_OUTFILE = ROOT / "charting-react" / "public" / "data" / "fundamentals" / "deliveries.json"

SHEET = "Sheet1"

COMMODITY_COLUMNS = {
    "White Maize": {
        "weekly": 7,
        "cec": "WM",
    },
    "Yellow Maize": {
        "weekly": 10,
        "cec": "YM",
    },
    "Total Maize": {
        "weekly": 13,
        "cec": "TOTAL",
    },
}

GRADE_COLUMNS = {
    "White Maize": {
        "WM1": 14,
        "WM2": 15,
        "WM3": 16,
        "WMO": 17,
        "White Total": 18,
    },
    "Yellow Maize": {
        "YM1": 19,
        "YM2": 20,
        "YM3": 21,
        "YMO": 22,
        "Yellow Total": 23,
    },
    "Total Maize": {
        "White Total": 18,
        "Yellow Total": 23,
        "Grand Total": 24,
    },
}


def as_number(value: object) -> float:
    if value in (None, "", "n/a", "#DIV/0!"):
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def valid_year(value: object) -> bool:
    return isinstance(value, str) and "/" in value and not value.startswith("#")


def load_cec(ws) -> dict[str, dict[str, float]]:
    estimates: dict[str, dict[str, float]] = {}
    for row in ws.iter_rows(min_row=3, min_col=29, max_col=32, values_only=True):
        year, wm, ym, total = row
        if valid_year(year):
            estimates[str(year)] = {
                "White Maize": as_number(wm),
                "Yellow Maize": as_number(ym),
                "Total Maize": as_number(total),
            }
    return estimates


def load_oilseed_cec(ws, commodity: str) -> dict[str, dict[str, float]]:
    estimates: dict[str, dict[str, float]] = {}
    for year, estimate in ws.iter_rows(min_row=2, max_row=12, min_col=11, max_col=12, values_only=True):
        if valid_year(year):
            estimates[str(year)] = {commodity: as_number(estimate)}
    return estimates


def load_oilseed_rows() -> tuple[list[dict], dict[str, dict[str, float]], set[str]]:
    if not OILSEEDS_WORKBOOK.exists():
        return [], {}, set()

    wb = load_workbook(OILSEEDS_WORKBOOK, data_only=True, read_only=True)
    sheet_map = {"Sunflowers": "Sunflowers", "Soybean": "Soybeans"}
    rows: list[dict] = []
    cec: dict[str, dict[str, float]] = defaultdict(dict)
    years: set[str] = set()

    for sheet_name, commodity in sheet_map.items():
        ws = wb[sheet_name]
        commodity_cec = load_oilseed_cec(ws, commodity)
        for year, values in commodity_cec.items():
            cec[year].update(values)

        running: dict[str, float] = defaultdict(float)
        for excel_row in ws.iter_rows(min_row=2, min_col=1, max_col=6, values_only=True):
            week, year, week_label, _prod, _adjustments, weekly = excel_row
            if not valid_year(year) or not isinstance(week, (int, float)):
                continue
            week = int(week)
            if week > 52:
                continue

            year = str(year)
            years.add(year)
            running[year] += as_number(weekly)
            cec_value = cec.get(year, {}).get(commodity, 0.0)
            cumulative = running[year]
            rows.append(
                {
                    "family": "Oilseeds",
                    "methodology": "Standard",
                    "marketingYear": year,
                    "weekNumber": week,
                    "weekLabel": str(week_label or ""),
                    "commodity": commodity,
                    "weeklyTons": round(as_number(weekly), 3),
                    "cumulativeTons": round(cumulative, 3),
                    "cecEstimate": round(cec_value, 3),
                    "percentDelivered": round((cumulative / cec_value) * 100, 6) if cec_value else None,
                }
            )

    return rows, dict(cec), years


def normalise() -> dict:
    wb = load_workbook(WORKBOOK, data_only=True, read_only=True)
    ws = wb[SHEET]
    cec = load_cec(ws)

    delivery_rows: list[dict] = []
    grade_weekly: dict[tuple[str, str, int, str], float] = defaultdict(float)
    grade_labels = {commodity: list(grades.keys()) for commodity, grades in GRADE_COLUMNS.items()}
    running_deliveries: dict[tuple[str, str, str], float] = defaultdict(float)
    years: set[str] = set()

    for excel_row in ws.iter_rows(min_row=2, min_col=1, max_col=25, values_only=True):
        sagis_week = excel_row[0]
        earlies_week = excel_row[1]
        week_label = str(excel_row[2] or "")
        sagis_year = excel_row[3]
        earlies_year = excel_row[4]
        if not week_label:
            continue

        for methodology, year, week in (
            ("SAGIS", sagis_year, sagis_week),
            ("Earlies", earlies_year, earlies_week),
        ):
            if not valid_year(year) or not isinstance(week, (int, float)):
                continue
            year = str(year)
            week = int(week)
            if week > 52:
                continue
            years.add(year)

            for commodity, indexes in COMMODITY_COLUMNS.items():
                weekly = as_number(excel_row[indexes["weekly"]])
                running_key = (methodology, year, commodity)
                running_deliveries[running_key] += weekly
                cumulative = running_deliveries[running_key]
                cec_value = cec.get(year, {}).get(commodity, 0.0)
                delivery_rows.append(
                    {
                        "family": "Maize",
                        "methodology": methodology,
                        "marketingYear": year,
                        "weekNumber": week,
                        "weekLabel": week_label,
                        "commodity": commodity,
                        "weeklyTons": round(weekly, 3),
                        "cumulativeTons": round(cumulative, 3),
                        "cecEstimate": round(cec_value, 3),
                        "percentDelivered": round((cumulative / cec_value) * 100, 6) if cec_value else None,
                    }
                )

            # Grade charts compare weekly grade deliveries across marketing years.
            # The methodology changes which week/year bucket the same weekly data belongs to.
            for commodity, grades in GRADE_COLUMNS.items():
                for grade, col_idx in grades.items():
                    grade_weekly[(methodology, year, week, commodity, grade)] += as_number(excel_row[col_idx])

    grade_rows: list[dict] = []
    running: dict[tuple[str, str, str, str], float] = defaultdict(float)
    for key in sorted(grade_weekly, key=lambda item: (item[0], item[3], item[4], item[1], item[2])):
        methodology, year, week, commodity, grade = key
        weekly = grade_weekly[key]
        running_key = (methodology, year, commodity, grade)
        running[running_key] += weekly
        total_grade = "Grand Total" if commodity == "Total Maize" else "White Total" if commodity == "White Maize" else "Yellow Total"
        total_key = (methodology, year, week, commodity, total_grade)
        total_weekly = grade_weekly.get(total_key, 0.0)
        grade_rows.append(
            {
                "methodology": methodology,
                "marketingYear": year,
                "weekNumber": week,
                "commodity": commodity,
                "grade": grade,
                "weeklyTons": round(weekly, 3),
                "cumulativeTons": round(running[running_key], 3),
                "totalWeeklyTons": round(total_weekly, 3),
            }
        )

    total_cumulative_lookup = {
        (row["methodology"], row["marketingYear"], row["weekNumber"], row["commodity"]): row["cumulativeTons"]
        for row in grade_rows
        if row["grade"] in {"White Total", "Yellow Total", "Grand Total"}
    }
    for row in grade_rows:
        total = total_cumulative_lookup.get((row["methodology"], row["marketingYear"], row["weekNumber"], row["commodity"]), 0.0)
        row["percentOfTotalDelivered"] = round((row["cumulativeTons"] / total) * 100, 6) if total else None

    delivery_rows.sort(key=lambda item: (item["methodology"], item["commodity"], item["marketingYear"], item["weekNumber"]))
    grade_rows.sort(key=lambda item: (item["methodology"], item["commodity"], item["grade"], item["marketingYear"], item["weekNumber"]))
    oilseed_rows, oilseed_cec, oilseed_years = load_oilseed_rows()
    delivery_rows.extend(oilseed_rows)
    for year, values in oilseed_cec.items():
        cec.setdefault(year, {}).update(values)
    years.update(oilseed_years)
    delivery_rows.sort(key=lambda item: (item["family"], item["methodology"], item["commodity"], item["marketingYear"], item["weekNumber"]))

    return {
        "generatedFrom": WORKBOOK.name,
        "oilseedsGeneratedFrom": OILSEEDS_WORKBOOK.name if OILSEEDS_WORKBOOK.exists() else None,
        "families": ["Maize", "Sunflowers", "Soybeans"],
        "methodologies": ["SAGIS", "Earlies"],
        "commodities": ["White Maize", "Yellow Maize", "Total Maize"],
        "oilseedCommodities": ["Sunflowers", "Soybeans"],
        "marketingYears": sorted(years),
        "cecEstimates": cec,
        "gradeOptions": grade_labels,
        "deliveryRows": delivery_rows,
        "gradeRows": grade_rows,
    }


def main() -> None:
    OUTFILE.parent.mkdir(exist_ok=True)
    payload = json.dumps(normalise(), indent=2)
    OUTFILE.write_text(payload, encoding="utf-8")
    PUBLIC_OUTFILE.parent.mkdir(parents=True, exist_ok=True)
    PUBLIC_OUTFILE.write_text(payload, encoding="utf-8")
    print(f"Wrote {OUTFILE}")
    print(f"Wrote {PUBLIC_OUTFILE}")


if __name__ == "__main__":
    main()
